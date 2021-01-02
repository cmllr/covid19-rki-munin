#!/usr/bin/env python3

from sys import argv
from datetime import datetime
from os.path import join
from openpyxl import load_workbook

states = [
    "Baden-Württemberg",
    "Bayern",
    "Berlin",
    "Brandenburg",
    "Bremen",
    "Hamburg",
    "Hessen",
    "Mecklenburg-Vorpommern",
    "Niedersachsen",
    "Nordrhein-Westfalen",
    "Rheinland-Pfalz",
    "Saarland",
    "Sachsen",
    "Sachsen-Anhalt",
    "Schleswig-Holstein",
    "Thüringen"
]

EXCEL_DUMP_PATH = "/path/where/the/files/are/located"


def output_config():
    print("graph_title COVID-19: Vaccinations in DE")
    print("graph_category COVID-19")
    print("graph_info Based on RKI info, graph updates daily at 3pm")
    print("graph_scale no")
    for index, state in enumerate(states):
        print("de_{}.label {}".format(index, state))


def output_values():
    now = datetime.now()
    date_string = now.strftime("%d.%m.%y")

    files = list(filter(isfile, glob.glob(EXCEL_DUMP_PATH + "*.xlsx")))
    files.sort(key=lambda x: getmtime(x))
    if len(files) > 0:
        file_name = files[-1]
        wb = load_workbook(file_name)
        ws = wb[wb.sheetnames[1]]
        for index, state in enumerate(states):
            vaccinations = ws["B{}".format(index + 2)].value
            print("de_{}.value {}".format(index, vaccinations))


if len(argv) == 2 and argv[1] == "config":
    output_config()
else:
    output_values()
